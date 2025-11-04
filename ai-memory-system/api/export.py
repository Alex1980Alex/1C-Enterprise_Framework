"""
Export Module
Экспорт результатов поиска и истории в различные форматы
"""

import csv
import json
import io
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

# Опциональный импорт для Excel
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel export disabled")


class SearchResultsExporter:
    """Экспорт результатов поиска в различные форматы"""

    @staticmethod
    def to_csv(results: List[Dict[str, Any]], query: str = "") -> str:
        """
        Экспорт результатов в CSV формат

        Args:
            results: Список результатов поиска
            query: Поисковый запрос (для метаданных)

        Returns:
            CSV строка
        """
        if not results:
            return "No results to export"

        output = io.StringIO()

        # Определяем поля для экспорта
        fieldnames = [
            'file_path',
            'module_type',
            'score',
            'functions_count',
            'variables_count',
            'summary'
        ]

        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')

        # Заголовок с метаданными
        output.write(f"# BSL Code Search Results\n")
        output.write(f"# Query: {query}\n")
        output.write(f"# Exported: {datetime.now().isoformat()}\n")
        output.write(f"# Total Results: {len(results)}\n")
        output.write("\n")

        writer.writeheader()

        for result in results:
            # Подготовка данных для CSV
            row = {
                'file_path': result.get('file_path', ''),
                'module_type': result.get('module_type', ''),
                'score': f"{result.get('score', 0.0):.4f}",
                'functions_count': result.get('functions_count', 0),
                'variables_count': result.get('variables_count', 0),
                'summary': result.get('summary', '')[:200]  # Ограничение длины
            }
            writer.writerow(row)

        content = output.getvalue()
        output.close()

        logger.info(f"Exported {len(results)} results to CSV format")
        return content

    @staticmethod
    def to_json(results: List[Dict[str, Any]], query: str = "",
                include_code: bool = False) -> str:
        """
        Экспорт результатов в JSON формат

        Args:
            results: Список результатов поиска
            query: Поисковый запрос
            include_code: Включать ли код функций/процедур

        Returns:
            JSON строка
        """
        export_data = {
            "metadata": {
                "query": query,
                "exported_at": datetime.now().isoformat(),
                "total_results": len(results),
                "include_code": include_code
            },
            "results": []
        }

        for result in results:
            # Базовые поля
            export_result = {
                "file_path": result.get('file_path'),
                "module_type": result.get('module_type'),
                "score": result.get('score'),
                "functions_count": result.get('functions_count'),
                "variables_count": result.get('variables_count'),
                "summary": result.get('summary')
            }

            # Опционально включаем код
            if include_code:
                export_result['functions'] = result.get('functions', [])
                export_result['procedures'] = result.get('procedures', [])
            else:
                # Только названия функций/процедур
                export_result['function_names'] = [
                    f.get('name') for f in result.get('functions', [])
                ]
                export_result['procedure_names'] = [
                    p.get('name') for p in result.get('procedures', [])
                ]

            export_data["results"].append(export_result)

        content = json.dumps(export_data, ensure_ascii=False, indent=2)
        logger.info(f"Exported {len(results)} results to JSON format")
        return content

    @staticmethod
    def to_excel(results: List[Dict[str, Any]], query: str = "") -> bytes:
        """
        Экспорт результатов в Excel формат

        Args:
            results: Список результатов поиска
            query: Поисковый запрос

        Returns:
            Байты Excel файла

        Raises:
            RuntimeError: Если openpyxl не установлен
        """
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl library is required for Excel export")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Search Results"

        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Заголовок документа
        ws['A1'] = f"BSL Code Search Results - Query: {query}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')

        ws['A2'] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:F2')

        ws['A3'] = f"Total Results: {len(results)}"
        ws.merge_cells('A3:F3')

        # Заголовки колонок
        headers = ['File Path', 'Module Type', 'Score', 'Functions', 'Variables', 'Summary']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Данные
        for row_idx, result in enumerate(results, start=6):
            ws.cell(row=row_idx, column=1, value=result.get('file_path', ''))
            ws.cell(row=row_idx, column=2, value=result.get('module_type', ''))
            ws.cell(row=row_idx, column=3, value=result.get('score', 0.0))
            ws.cell(row=row_idx, column=4, value=result.get('functions_count', 0))
            ws.cell(row=row_idx, column=5, value=result.get('variables_count', 0))
            ws.cell(row=row_idx, column=6, value=result.get('summary', '')[:500])

        # Автоширина колонок
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
        ws.column_dimensions['A'].width = 50  # File path
        ws.column_dimensions['F'].width = 60  # Summary

        # Сохранение в байты
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        content = output.read()
        output.close()

        logger.info(f"Exported {len(results)} results to Excel format")
        return content


class HistoryExporter:
    """Экспорт истории поиска в различные форматы"""

    @staticmethod
    def to_csv(history_entries: List[Dict[str, Any]]) -> str:
        """
        Экспорт истории в CSV формат

        Args:
            history_entries: Список записей истории

        Returns:
            CSV строка
        """
        if not history_entries:
            return "No history to export"

        output = io.StringIO()

        fieldnames = [
            'id',
            'timestamp',
            'query',
            'results_count',
            'search_time_ms',
            'filters',
            'user_id'
        ]

        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')

        # Заголовок
        output.write(f"# BSL Code Search History\n")
        output.write(f"# Exported: {datetime.now().isoformat()}\n")
        output.write(f"# Total Entries: {len(history_entries)}\n")
        output.write("\n")

        writer.writeheader()

        for entry in history_entries:
            row = {
                'id': entry.get('id'),
                'timestamp': entry.get('timestamp'),
                'query': entry.get('query'),
                'results_count': entry.get('results_count'),
                'search_time_ms': f"{entry.get('search_time_ms', 0.0):.2f}",
                'filters': json.dumps(entry.get('filters')) if entry.get('filters') else '',
                'user_id': entry.get('user_id', '')
            }
            writer.writerow(row)

        content = output.getvalue()
        output.close()

        logger.info(f"Exported {len(history_entries)} history entries to CSV")
        return content

    @staticmethod
    def to_json(history_entries: List[Dict[str, Any]]) -> str:
        """
        Экспорт истории в JSON формат

        Args:
            history_entries: Список записей истории

        Returns:
            JSON строка
        """
        export_data = {
            "metadata": {
                "exported_at": datetime.now().isoformat(),
                "total_entries": len(history_entries)
            },
            "history": history_entries
        }

        content = json.dumps(export_data, ensure_ascii=False, indent=2)
        logger.info(f"Exported {len(history_entries)} history entries to JSON")
        return content

    @staticmethod
    def to_excel(history_entries: List[Dict[str, Any]]) -> bytes:
        """
        Экспорт истории в Excel формат

        Args:
            history_entries: Список записей истории

        Returns:
            Байты Excel файла

        Raises:
            RuntimeError: Если openpyxl не установлен
        """
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl library is required for Excel export")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Search History"

        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Заголовок
        ws['A1'] = "BSL Code Search History"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:G1')

        ws['A2'] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:G2')

        ws['A3'] = f"Total Entries: {len(history_entries)}"
        ws.merge_cells('A3:G3')

        # Заголовки колонок
        headers = ['ID', 'Timestamp', 'Query', 'Results', 'Time (ms)', 'Filters', 'User ID']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill

        # Данные
        for row_idx, entry in enumerate(history_entries, start=6):
            ws.cell(row=row_idx, column=1, value=entry.get('id'))
            ws.cell(row=row_idx, column=2, value=entry.get('timestamp'))
            ws.cell(row=row_idx, column=3, value=entry.get('query'))
            ws.cell(row=row_idx, column=4, value=entry.get('results_count'))
            ws.cell(row=row_idx, column=5, value=entry.get('search_time_ms'))

            filters = entry.get('filters')
            ws.cell(row=row_idx, column=6, value=json.dumps(filters) if filters else '')
            ws.cell(row=row_idx, column=7, value=entry.get('user_id', ''))

        # Автоширина
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 15
        ws.column_dimensions['C'].width = 40  # Query
        ws.column_dimensions['F'].width = 30  # Filters

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        content = output.read()
        output.close()

        logger.info(f"Exported {len(history_entries)} history entries to Excel")
        return content
